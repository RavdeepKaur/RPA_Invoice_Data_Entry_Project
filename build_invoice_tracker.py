import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "extracted_invoice_data.csv"
OUTPUT_XLSX = ROOT / "Invoice_Data_Entry_Tracker.xlsx"


def main():
    with CSV_PATH.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    headers = [
        "source_file", "invoice_no", "vendor", "invoice_date", "due_date", "po_number",
        "subtotal", "tax", "total", "validation_status", "validation_notes",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top")

    for col in ["G", "H", "I"]:
        for cell in ws[col][1:]:
            cell.value = float(cell.value)
            cell.number_format = "$#,##0.00"

    widths = [18, 14, 28, 14, 14, 14, 12, 12, 12, 18, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws["M1"] = "Automation Summary"
    ws["M1"].fill = PatternFill("solid", fgColor="E8F1FA")
    ws["M1"].font = Font(bold=True, color="1F4E79", size=12)
    summary = [
        ("Invoices Processed", f"=COUNTA(B2:B{ws.max_row})"),
        ("Ready for Entry", f'=COUNTIF(J2:J{ws.max_row},"Ready for Entry")'),
        ("Needs Review", f'=COUNTIF(J2:J{ws.max_row},"Needs Review")'),
        ("Total Invoice Value", f"=SUM(I2:I{ws.max_row})"),
    ]
    for row_idx, (label, formula) in enumerate(summary, start=2):
        ws[f"M{row_idx}"] = label
        ws[f"N{row_idx}"] = formula
        ws[f"M{row_idx}"].font = Font(bold=True)
    ws["N5"].number_format = "$#,##0.00"
    ws.column_dimensions["M"].width = 24
    ws.column_dimensions["N"].width = 16

    ws["M7"] = "RPA Flow Logic"
    ws["M7"].fill = header_fill
    ws["M7"].font = header_font
    steps = [
        "Read invoice PDF files from folder",
        "Extract invoice fields using text patterns/OCR",
        "Validate required fields and total = subtotal + tax",
        "Write clean records to Excel and flag exceptions",
    ]
    for idx, step in enumerate(steps, start=8):
        ws[f"M{idx}"] = idx - 7
        ws[f"N{idx}"] = step
    ws.column_dimensions["N"].width = 55

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_XLSX)
    print(f"Created {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
